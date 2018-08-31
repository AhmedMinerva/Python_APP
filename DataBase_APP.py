import cv2
import os
import tkinter as tk
from openpyxl import Workbook, load_workbook
from tkinter import filedialog
import openpyxl

font = ("Verdana", 12)
ID = [1]

class Application(tk.Tk):
    
    def __init__(self):
        tk.Tk.__init__(self)
        self.title("CEA App")
        self.geometry("600x600")
        background_image=tk.PhotoImage(file="CEA_Logo.png")
        background_label = tk.Label(self, image=background_image)
        background_label.place(x=0, y=0, relwidth=1, relheight=1)
        background_label.pack()
        background_label.image = background_image
        
        container = tk.Frame(self)
        container.pack()
        
        self.frames = {}

        for F in (startpage, inspection_ID, page2, page3, page4, page5, page6, page7, page8, existing, existing2):
            frame = F(parent=container, controller=self)
            self.frames[F] = frame
            frame.grid(row=0, column=0, sticky="nsew") #stretch to all sides yemen shamal kol 7aga
        
        self.show(startpage)

    def show(self, page_name):
        frame = self.frames[page_name]
        frame.tkraise()
        val = inspection_ID.InspID.get()
        if val != "":
            ID[0] = inspection_ID.InspID.get()
        inspection_ID.InspID.delete(0, 1000)
        
        
       




class startpage(tk.Frame):
    
    def __init__(self, parent, controller):
        
        tk.Frame.__init__(self, parent) #-----------
        
        label1 = tk.Label(self, text = "Choose the Report: ", font = font)
        label1.pack(side="top", fill="x", pady=10)
        button1 =tk.Button(self, text= "IPM Report", command= lambda: controller.show(existing))
        button1.pack()
        button3 =tk.Button(self, text= "Flash Test Report", command= lambda: controller.show(page3))
        button3.place(x=0, y=0, relwidth=1, relheight=1)
        button3.pack()
        button = tk.Button(self, text = "Close", command = self.ex)
        button.pack()

        
        
    def ex(self):
        exit()
        

        

class existing(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        button1 = tk.Button(self, text = "New Report", command= lambda : controller.show(inspection_ID))
        button1.pack()
        button2 =tk.Button(self, text = "Existing Reports", command = lambda : controller.show(existing2) )
        button2.pack()
        button3 = tk.Button(self, text = "Go Back", command = lambda : controller.show(startpage))
        button3.pack()


class existing2(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        label = tk.Label(self, text = "Choose the report you want to edit:", font = font)
        label.pack()
        options = ["Select ▼"]
        wb = load_workbook("Report.xlsx")
        ws = wb.active
        first_column = ws['A']
        
        for x in range(1,len(first_column)):
            colll = first_column[x].value
            if colll not in options:
                 options.append(colll)

        variable = tk.StringVar(self)
        variable.set(options[0])
        drop = tk.OptionMenu(self, variable, *options)
        drop.pack()
        button = tk.Button(self, text = "Go Back", command = lambda : controller.show(existing))
        button.pack()
        
        
        
        


class inspection_ID(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        label = tk.Label(self, text= "Enter the Inpection ID", font = font)
        label.pack()
        inspection_ID.InspID = tk.Entry(self, text = "Enter the Inspection ID")
        inspection_ID.InspID.pack()
        button1 =tk.Button(self, text= "Enter", command= lambda: controller.show(page2))
        button1.pack()
        button3 = tk.Button(self, text = "Go Back", command = lambda : controller.show(existing))
        button3.pack()

        


class page2(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        label2 = tk.Label(self, text = "IPM Report")
        button4 =tk.Button(self, text= "go back", command= lambda: controller.show(startpage))
        label2.pack()
        button5 =tk.Button(self, text= "Flux", command= lambda: controller.show(page4))
        button6 =tk.Button(self, text= "Cell", command= lambda: controller.show(page5))
        button7 =tk.Button(self, text= "Junction box", command= lambda: controller.show(page6))
        button8 =tk.Button(self, text= "Silicone", command= lambda: controller.show(page7))
        button9 =tk.Button(self, text= "Glass", command= lambda: controller.show(page8))
        button5.pack()
        button6.pack()
        button7.pack()
        button8.pack()
        button9.pack()
        button4.pack()
        
class page3(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        
        label3 = tk.Label(self, text = "Flash Test Report")
        button5 =tk.Button(self, text= "go back", command= lambda: controller.show(startpage))
        label3.pack()
        button5.pack()        
        

class page4(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        
        self.label3 = tk.Label(self, text = "Flux")
        button5 =tk.Button(self, text= "go back", command= lambda: controller.show(page2))
        self.label3.pack()
        
        
        self.dlabel1 = tk.Label(self, text = "material supplier")
        self.data1 = tk.Entry(self)
        self.dlabel1.pack()
        self.data1.pack()
        
        self.dlabel2 = tk.Label(self, text = "Material supplier number")
        self.data2 = tk.Entry(self)
        self.dlabel2.pack()
        self.data2.pack()
        
        self.dlabel3 = tk.Label(self, text = "Type")
        self.data3 = tk.Entry(self)
        self.dlabel3.pack()
        self.data3.pack()
        
        self.dlabel4 = tk.Label(self, text = "Resolution")
        self.data4 = tk.Entry(self)
        self.dlabel4.pack()
        self.data4.pack()
        
        self.dlabel5 = tk.Label(self, text = "Expiration Date")
        self.data5 = tk.Entry(self)
        self.dlabel5.pack()
        self.data5.pack()
        self.buttonSave = tk.Button(self, text = "Save", command = self.save)
        self.buttonSave.pack()
        buttonn = tk.Button(self, text = "browse", command = self.im)
        buttonn.pack()
        button5.pack()


        
        
        self.x = 0
    def im(self):
        self.x = 1
        self.filename =  filedialog.askopenfilename(initialdir = "/",title = "Select file",filetypes = (("jpeg files","*.jpg"),("all files","*.*")))
    


    def save(self):
        saving = []
        saving.append(ID[0])
        

        saving.append("Flux")
        
        data1 = self.data1.get()
        saving.append(data1)
        self.data1.delete(0, 1000)
        
        data2 = self.data2.get()
        saving.append(data2)
        self.data2.delete(0, 1000)
        
        data3 = self.data3.get()
        saving.append(data3)
        self.data3.delete(0, 1000)
        
        data4 = self.data4.get()
        saving.append(data4)
        self.data4.delete(0, 1000)
        
        data5 = self.data5.get()
        saving.append(data5)
        self.data5.delete(0, 1000)
        
        wb = load_workbook("Report.xlsx")
        ws = wb.active
        ws.append(saving)
        if self.x == 1:
            
            if len(self.filename) != 0:

                img = cv2.imread(self.filename, 1)
                path = r"C:\Users\gaber\AppData\Local\Programs\Python\Python36-32\Pictures"
                cv2.imwrite(os.path.join(path , "Flux.jpg"), img)
                cv2.waitKey(0)
                
                wb.save("Report.xlsx")
                
        else:
            wb.save("Report.xlsx")
        


class page5(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        
        label3 = tk.Label(self, text = "Cell")
        button5 =tk.Button(self, text= "go back", command= lambda: controller.show(page2))
        label3.pack()
        
        self.dlabel1 = tk.Label(self, text = "material supplier")
        self.data1 = tk.Entry(self)
        self.dlabel1.pack()
        self.data1.pack()
        
        self.dlabel2 = tk.Label(self, text = "Material supplier number")
        self.data2 = tk.Entry(self)
        self.dlabel2.pack()
        self.data2.pack()
        
        self.dlabel3 = tk.Label(self, text = "Type")
        self.data3 = tk.Entry(self)
        self.dlabel3.pack()
        self.data3.pack()
        
        self.dlabel4 = tk.Label(self, text = "Resolution")
        self.data4 = tk.Entry(self)
        self.dlabel4.pack()
        self.data4.pack()
        
        self.dlabel5 = tk.Label(self, text = "Expiration Date")
        self.data5 = tk.Entry(self)
        self.dlabel5.pack()
        self.data5.pack()
        self.buttonSave = tk.Button(self, text = "Save", command = self.save)
        self.buttonSave.pack()

        buttonn = tk.Button(self, text = "browse", command = self.im)
        buttonn.pack()
        button5.pack()

           
        self.x = 0
    def im(self):
        self.x = 1
        self.filename =  filedialog.askopenfilename(initialdir = "/",title = "Select file",filetypes = (("jpeg files","*.jpg"),("all files","*.*")))
    


    def save(self):
        saving = []
        saving.append(ID[0])

        saving.append("Cell")
        
        data1 = self.data1.get()
        saving.append(data1)
        self.data1.delete(0, 1000)
        
        data2 = self.data2.get()
        saving.append(data2)
        self.data2.delete(0, 1000)
        
        data3 = self.data3.get()
        saving.append(data3)
        self.data3.delete(0, 1000)
        
        data4 = self.data4.get()
        saving.append(data4)
        self.data4.delete(0, 1000)
        
        data5 = self.data5.get()
        saving.append(data5)
        self.data5.delete(0, 1000)
        
        wb = load_workbook("Report.xlsx")
        ws = wb.active
        ws.append(saving)
        if self.x == 1:
            
            if len(self.filename) != 0:

                img = cv2.imread(self.filename, 1)
                path = r"C:\Users\gaber\AppData\Local\Programs\Python\Python36-32\Pictures"
                cv2.imwrite(os.path.join(path , "Cell.jpg"), img)
                cv2.waitKey(0)
                
                wb.save("Report.xlsx")
                
        else:
            wb.save("Report.xlsx")
        
class page6(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        
        label3 = tk.Label(self, text = "Junction Box")
        button5 =tk.Button(self, text= "go back", command= lambda: controller.show(page2))
        label3.pack()
        
                
        self.dlabel1 = tk.Label(self, text = "material supplier")
        self.data1 = tk.Entry(self)
        self.dlabel1.pack()
        self.data1.pack()
        
        self.dlabel2 = tk.Label(self, text = "Material supplier number")
        self.data2 = tk.Entry(self)
        self.dlabel2.pack()
        self.data2.pack()
        
        self.dlabel3 = tk.Label(self, text = "Type")
        self.data3 = tk.Entry(self)
        self.dlabel3.pack()
        self.data3.pack()
        
        self.dlabel4 = tk.Label(self, text = "Resolution")
        self.data4 = tk.Entry(self)
        self.dlabel4.pack()
        self.data4.pack()
        
        self.dlabel5 = tk.Label(self, text = "Expiration Date")
        self.data5 = tk.Entry(self)
        self.dlabel5.pack()
        self.data5.pack()
        self.buttonSave = tk.Button(self, text = "Save", command = self.save)
        self.buttonSave.pack()


        buttonn = tk.Button(self, text = "browse", command = self.im)
        buttonn.pack()
        button5.pack()

           
        self.x = 0
    def im(self):
        self.x = 1
        self.filename =  filedialog.askopenfilename(initialdir = "/",title = "Select file",filetypes = (("jpeg files","*.jpg"),("all files","*.*")))
    


    def save(self):
            saving = []
            saving.append(ID[0])

            saving.append("Junction Box")
            
            data1 = self.data1.get()
            saving.append(data1)
            self.data1.delete(0, 1000)
            
            data2 = self.data2.get()
            saving.append(data2)
            self.data2.delete(0, 1000)
            
            data3 = self.data3.get()
            saving.append(data3)
            self.data3.delete(0, 1000)
            
            data4 = self.data4.get()
            saving.append(data4)
            self.data4.delete(0, 1000)
            
            data5 = self.data5.get()
            saving.append(data5)
            self.data5.delete(0, 1000)
            
            wb = load_workbook("Report.xlsx")
            ws = wb.active
            ws.append(saving)
            if self.x == 1:
                
                if len(self.filename) != 0:

                    img = cv2.imread(self.filename, 1)
                    path = r"C:\Users\gaber\AppData\Local\Programs\Python\Python36-32\Pictures"
                    cv2.imwrite(os.path.join(path , "Junction Box.jpg"), img)
                    cv2.waitKey(0)
                    
                    wb.save("Report.xlsx")
                    
            else:
                wb.save("Report.xlsx")
       

class page7(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        
        label3 = tk.Label(self, text = "Silicone")
        button5 =tk.Button(self, text= "go back", command= lambda: controller.show(page2))
        label3.pack()
                
        self.dlabel1 = tk.Label(self, text = "material supplier")
        self.data1 = tk.Entry(self)
        self.dlabel1.pack()
        self.data1.pack()
        
        self.dlabel2 = tk.Label(self, text = "Material supplier number")
        self.data2 = tk.Entry(self)
        self.dlabel2.pack()
        self.data2.pack()
        
        self.dlabel3 = tk.Label(self, text = "Type")
        self.data3 = tk.Entry(self)
        self.dlabel3.pack()
        self.data3.pack()
        
        self.dlabel4 = tk.Label(self, text = "Resolution")
        self.data4 = tk.Entry(self)
        self.dlabel4.pack()
        self.data4.pack()
        
        self.dlabel5 = tk.Label(self, text = "Expiration Date")
        self.data5 = tk.Entry(self)
        self.dlabel5.pack()
        self.data5.pack()
        self.buttonSave = tk.Button(self, text = "Save", command = self.save)
        self.buttonSave.pack()


        buttonn = tk.Button(self, text = "browse", command = self.im)
        buttonn.pack()
        button5.pack()

           
        self.x = 0
    def im(self):
        self.x = 1
        self.filename =  filedialog.askopenfilename(initialdir = "/",title = "Select file",filetypes = (("jpeg files","*.jpg"),("all files","*.*")))
    


    def save(self):
        saving = []
        saving.append(ID[0])

        saving.append("Silicone")
        
        data1 = self.data1.get()
        saving.append(data1)
        self.data1.delete(0, 1000)
        
        data2 = self.data2.get()
        saving.append(data2)
        self.data2.delete(0, 1000)
        
        data3 = self.data3.get()
        saving.append(data3)
        self.data3.delete(0, 1000)
        
        data4 = self.data4.get()
        saving.append(data4)
        self.data4.delete(0, 1000)
        
        data5 = self.data5.get()
        saving.append(data5)
        self.data5.delete(0, 1000)
        
        wb = load_workbook("Report.xlsx")
        ws = wb.active
        ws.append(saving)
        if self.x == 1:
            
            if len(self.filename) != 0:

                img = cv2.imread(self.filename, 1)
                path = r"C:\Users\gaber\AppData\Local\Programs\Python\Python36-32\Pictures"
                cv2.imwrite(os.path.join(path , "Silicone.jpg"), img)
                cv2.waitKey(0)
                
                wb.save("Report.xlsx")
                
        else:
            wb.save("Report.xlsx")

class page8(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        
        label3 = tk.Label(self, text = "Glass")
        button5 =tk.Button(self, text= "go back", command= lambda: controller.show(page2))
        label3.pack()
        
                
        self.dlabel1 = tk.Label(self, text = "material supplier")
        self.data1 = tk.Entry(self)
        self.dlabel1.pack()
        self.data1.pack()
        
        self.dlabel2 = tk.Label(self, text = "Material supplier number")
        self.data2 = tk.Entry(self)
        self.dlabel2.pack()
        self.data2.pack()
        
        self.dlabel3 = tk.Label(self, text = "Type")
        self.data3 = tk.Entry(self)
        self.dlabel3.pack()
        self.data3.pack()
        
        self.dlabel4 = tk.Label(self, text = "Resolution")
        self.data4 = tk.Entry(self)
        self.dlabel4.pack()
        self.data4.pack()
        
        self.dlabel5 = tk.Label(self, text = "Expiration Date")
        self.data5 = tk.Entry(self)
        self.dlabel5.pack()
        self.data5.pack()
        self.buttonSave = tk.Button(self, text = "Save", command = self.save)
        self.buttonSave.pack()


        buttonn = tk.Button(self, text = "browse", command = self.im)
        buttonn.pack()
        button5.pack()

           
        self.x = 0
    def im(self):
        self.x = 1
        self.filename =  filedialog.askopenfilename(initialdir = "/",title = "Select file",filetypes = (("jpeg files","*.jpg"),("all files","*.*")))
    


    def save(self):
        saving = []
        saving.append(ID[0])

        saving.append("Glass")
        
        data1 = self.data1.get()
        saving.append(data1)
        self.data1.delete(0, 1000)
        
        data2 = self.data2.get()
        saving.append(data2)
        self.data2.delete(0, 1000)
        
        data3 = self.data3.get()
        saving.append(data3)
        self.data3.delete(0, 1000)
        
        data4 = self.data4.get()
        saving.append(data4)
        self.data4.delete(0, 1000)
        
        data5 = self.data5.get()
        saving.append(data5)
        self.data5.delete(0, 1000)
        
        wb = load_workbook("Report.xlsx")
        ws = wb.active
        ws.append(saving)
        if self.x == 1:
            
            if len(self.filename) != 0:

                img = cv2.imread(self.filename, 1)
                path = r"C:\Users\gaber\AppData\Local\Programs\Python\Python36-32\Pictures"
                cv2.imwrite(os.path.join(path , "Glass.jpg"), img)
                cv2.waitKey(0)
                
                wb.save("Report.xlsx")
                
        else:
            wb.save("Report.xlsx")

            
if __name__ == "__main__":
    app = Application()
    app.mainloop()
        
            
